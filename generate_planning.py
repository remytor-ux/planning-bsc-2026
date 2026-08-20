#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Génère une page HTML (index.html) reproduisant fidèlement la mise en page du
fichier BSC_Planning_2026V2.xlsx : grille journalière + tableau horaire récap,
un onglet par mois, en lecture seule.

Usage :
    python3 generate_planning_v2.py chemin/vers/BSC_Planning_2026V2.xlsx dossier_sortie/
"""

import sys
import json
import openpyxl
from pathlib import Path
from datetime import datetime

MOIS_ORDRE = [
    "JANVIER", "FEVRIER", "MARS", "AVRIL", "MAI", "JUIN",
    "JUILLET", "AOUT", "SEPTEMBRE", "OCTOBRE", "NOVEMBRE", "DECEMBRE",
]

# Couleur par mot-clé détecté dans le libellé de colonne (palette navy/amber/red/blue, pas de vert)
COLOR_RULES = [
    ("19H00-9H00", "#8B1E2B"),   # nuit -> rouge
    ("19H-9H", "#8B1E2B"),
    ("9H00-21H00", "#1F3A5F"),   # 12h -> navy
    ("9H21", "#1F3A5F"),
    ("9H00-19H00", "#C99A3B"),   # jour -> amber
    ("9H-19H", "#C99A3B"),
    ("FFI", "#3B6EA5"),          # bleu
    ("UHCD", "#5A4B81"),         # violet
    ("RUF", "#2E6E6E"),          # sarcelle sobre
    ("DPI", "#6B5B3E"),          # brun
    ("USC", "#6B5B3E"),          # brun (même famille que DPI, poste équivalent)
    ("CA", "#7A7A7A"),           # gris
    ("INDISPO", "#B0413E"),
    ("FORMATION", "#4A5D8A"),
    ("AT", "#A0522D"),
]
VACANT_COLOR = "#8B1E2B"  # rouge : poste vacant (cellule rouge sans nom dans le fichier source)


def color_for_label(label):
    up = label.upper().replace("\n", " ")
    for key, color in COLOR_RULES:
        if key in up:
            return color
    return "#555555"


def is_red_fill(cell):
    """Détecte une cellule remplie en rouge (poste vacant) dans le fichier source."""
    fg = cell.fill.fgColor
    try:
        if fg.type == "rgb" and fg.rgb and str(fg.rgb).endswith("FF0000"):
            return True
    except Exception:
        pass
    return False


def parse_month_sheet(ws):
    # --- Colonnes journalières : lues dynamiquement depuis la ligne d'en-têtes (row 3)
    # et les regroupements de la ligne 2 (LIGNE 1 / 12H / LIGNE 2 / AUTRES / UHCD...),
    # car la structure varie légèrement d'un mois à l'autre dans le fichier source.
    group_row = [ws.cell(row=2, column=c).value for c in range(1, 26)]
    label_row = [ws.cell(row=3, column=c).value for c in range(1, 26)]

    # Reconstruit le groupe courant pour chaque colonne (les groupes sont fusionnés,
    # donc seule la première colonne du groupe porte la valeur -> on la propage).
    current_group = None
    groups = []
    for v in group_row:
        if v:
            current_group = str(v).strip()
        groups.append(current_group)

    day_columns = []
    for idx in range(2, len(label_row)):  # colonnes C (index 2, 0-based) et suivantes
        label = label_row[idx]
        if not label:
            continue
        col = idx + 1  # 1-based pour openpyxl
        label_clean = str(label).replace("\n", " ").strip()
        day_columns.append({
            "col": col,
            "label": label_clean,
            "group": groups[idx] or "AUTRES",
            "color": color_for_label(label_clean),
        })

    days = []
    r = 4
    while True:
        jour = ws.cell(row=r, column=1).value
        date_val = ws.cell(row=r, column=2).value
        if jour is None and date_val is None:
            break
        entry = {"jour": jour, "date": None, "cells": {}}
        if isinstance(date_val, datetime):
            entry["date"] = date_val.strftime("%d/%m")
        for dc in day_columns:
            cell = ws.cell(row=r, column=dc["col"])
            if cell.value:
                entry["cells"][dc["col"]] = {"val": str(cell.value).strip(), "vacant": False}
            elif is_red_fill(cell):
                entry["cells"][dc["col"]] = {"val": None, "vacant": True}
            else:
                entry["cells"][dc["col"]] = None
        days.append(entry)
        r += 1
        if r > 400:
            break

    # Tableau horaire récap : commence après la ligne "Nb semaines" (col C).
    # Les colonnes varient légèrement d'un onglet à l'autre (ex: DPI/USC selon
    # le mois), donc on lit les en-têtes réels ligne par ligne plutôt que de
    # supposer des indices de colonnes fixes.
    hours_start = None
    local_headers = []
    for rr in range(r, r + 40):
        if ws.cell(row=rr, column=3).value == "Nb semaines":
            header_row = rr + 1
            hours_start = rr + 2
            for col in range(4, 16):
                h = ws.cell(row=header_row, column=col).value
                if h:
                    local_headers.append((col, str(h).strip()))
            break
    hours = []
    if hours_start:
        rr = hours_start
        blanks_in_a_row = 0
        while blanks_in_a_row < 3:
            name = ws.cell(row=rr, column=3).value
            percent = ws.cell(row=rr, column=4).value
            if not name or not isinstance(name, str) or not isinstance(percent, (int, float)):
                blanks_in_a_row += 1
                rr += 1
                if rr - hours_start > 60:
                    break
                continue
            blanks_in_a_row = 0
            row = {"nom": str(name).strip()}
            for col, h in local_headers:
                v = ws.cell(row=rr, column=col).value
                if isinstance(v, float):
                    v = round(v, 1)
                row[h] = v
            hours.append(row)
            rr += 1
            if rr - hours_start > 60:
                break
    headers_out = ["Nom"] + [h for _, h in local_headers]
    return days, day_columns, hours, headers_out


def build_data(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    data = {}
    for name in MOIS_ORDRE:
        if name not in wb.sheetnames:
            continue
        days, day_columns, hours, hours_headers = parse_month_sheet(wb[name])
        data[name] = {
            "days": days,
            "columns": day_columns,
            "hours": hours,
            "hours_headers": hours_headers,
        }
    return data


HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Planning Urgences · SMUR · UHCD — CH Bagnols-sur-Cèze</title>
<style>
  :root {
    --navy: #1F3A5F;
    --amber: #C99A3B;
    --red: #8B1E2B;
    --blue: #3B6EA5;
    --bg: #F0F1F4;
    --card: #FFFFFF;
    --text: #1D2430;
    --muted: #6B7280;
    --border: #DADFE5;
  }
  * { box-sizing: border-box; }
  body { margin: 0; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Arial, sans-serif; background: var(--bg); color: var(--text); }
  header { background: var(--navy); color: #fff; padding: 14px 18px; position: sticky; top: 0; z-index: 20; box-shadow: 0 2px 6px rgba(0,0,0,0.15); }
  header h1 { margin: 0; font-size: 17px; font-weight: 600; }
  header .sub { font-size: 12px; opacity: 0.85; margin-top: 2px; }
  header .updated { font-size: 11px; opacity: 0.7; margin-top: 4px; }
  nav { display: flex; overflow-x: auto; background: #fff; border-bottom: 1px solid var(--border); position: sticky; top: 64px; z-index: 19; -webkit-overflow-scrolling: touch; }
  nav button { flex: 0 0 auto; padding: 9px 14px; background: none; border: none; font-size: 12.5px; font-weight: 600; color: var(--muted); cursor: pointer; border-bottom: 3px solid transparent; white-space: nowrap; }
  nav button.active { color: var(--navy); border-bottom-color: var(--amber); }
  main { padding: 10px; max-width: 100%; overflow-x: auto; }
  .month { display: none; }
  .month.active { display: block; }
  .table-wrap { overflow-x: auto; background: var(--card); border-radius: 8px; border: 1px solid var(--border); margin-bottom: 18px; }
  table { border-collapse: collapse; width: 100%; min-width: 900px; font-size: 12px; }
  thead th { background: var(--navy); color: #fff; padding: 6px 8px; text-align: center; font-weight: 600; border: 1px solid rgba(255,255,255,0.15); position: sticky; top: 0; }
  thead tr.group-row th { background: #16293f; font-size: 10.5px; letter-spacing: 0.03em; text-transform: uppercase; opacity: 0.9; }
  tbody td { padding: 5px 7px; border: 1px solid var(--border); text-align: center; white-space: nowrap; }
  tbody td.jour-cell { text-align: left; font-weight: 600; background: #FAFAFB; white-space: nowrap; }
  tbody td.date-cell { color: var(--muted); background: #FAFAFB; }
  tbody tr:nth-child(even) td:not(.jour-cell):not(.date-cell) { background: #FBFBFC; }
  .we td.jour-cell, .we td.date-cell { background: #EFEFF3; }
  .poste { display: inline-block; padding: 2px 6px; border-radius: 5px; color: #fff; font-weight: 600; font-size: 11.5px; }
  .poste.vacant { background: #1D2430; color: #fff; font-weight: 700; border: 2px solid #8B1E2B; }
  tr.row-vacant td { background: #FBEAEA !important; }
  tr.row-vacant td.jour-cell, tr.row-vacant td.date-cell { background: #F6D9D9 !important; box-shadow: inset 4px 0 0 #8B1E2B; }
  h2.month-title { font-size: 14px; color: var(--navy); margin: 4px 4px 8px; }
  h3.section-title { font-size: 13px; color: var(--navy); margin: 4px 4px 8px; }
  .hours-table td.nom { text-align: left; font-weight: 600; }
  .hours-table td.delta-pos { color: #1E6B3A; font-weight: 600; }
  .hours-table td.delta-neg { color: var(--red); font-weight: 600; }
  footer { text-align: center; font-size: 11px; color: var(--muted); padding: 18px; }
</style>
</head>
<body>
<header>
  <h1>Planning Urgences · SMUR · UHCD</h1>
  <div class="sub">CH Bagnols-sur-Cèze — lecture seule</div>
  <div class="updated">Dernière mise à jour : __GENERATED_AT__</div>
</header>
<nav id="monthNav"></nav>
<main id="monthContainer"></main>
<footer>Généré automatiquement depuis BSC_Planning_2026V2.xlsx — contactez Dr Torregrossa pour toute correction.</footer>

<script>
const DATA = __DATA_JSON__;
const HOURS_HEADERS = __HOURS_HEADERS_JSON__;

const nav = document.getElementById("monthNav");
const container = document.getElementById("monthContainer");
const months = Object.keys(DATA);

function groupSpans(dayColumns) {
  const spans = [];
  let last = null;
  dayColumns.forEach(c => {
    if (last && last.group === c.group) {
      last.span += 1;
    } else {
      last = {group: c.group, span: 1};
      spans.push(last);
    }
  });
  return spans;
}

months.forEach((month, idx) => {
  const btn = document.createElement("button");
  btn.textContent = month.charAt(0) + month.slice(1).toLowerCase();
  btn.dataset.month = month;
  if (idx === 0) btn.classList.add("active");
  btn.addEventListener("click", () => showMonth(month));
  nav.appendChild(btn);

  const section = document.createElement("div");
  section.className = "month" + (idx === 0 ? " active" : "");
  section.id = "month-" + month;

  // --- Grille journalière ---
  const dayColumns = DATA[month].columns;
  const wrap1 = document.createElement("div");
  wrap1.className = "table-wrap";
  const table = document.createElement("table");

  const thead = document.createElement("thead");
  const groupRow = document.createElement("tr");
  groupRow.className = "group-row";
  groupRow.innerHTML = "<th colspan='2'></th>";
  groupSpans(dayColumns).forEach(g => {
    groupRow.innerHTML += `<th colspan="${g.span}">${g.group}</th>`;
  });
  thead.appendChild(groupRow);

  const headRow = document.createElement("tr");
  headRow.innerHTML = "<th>Jour</th><th>Date</th>" + dayColumns.map(c => `<th>${c.label}</th>`).join("");
  thead.appendChild(headRow);
  table.appendChild(thead);

  const tbody = document.createElement("tbody");
  DATA[month].days.forEach(day => {
    const tr = document.createElement("tr");
    const hasVacant = Object.values(day.cells).some(c => c && c.vacant);
    if (hasVacant) {
      tr.className = "row-vacant";
    } else if (day.jour === "Samedi" || day.jour === "Dimanche") {
      tr.className = "we";
    }
    let rowHtml = `<td class="jour-cell">${day.jour || ""}</td><td class="date-cell">${day.date || ""}</td>`;
    dayColumns.forEach(c => {
      const cell = day.cells[c.col];
      if (cell && cell.val) {
        rowHtml += `<td><span class="poste" style="background:${c.color}">${cell.val}</span></td>`;
      } else if (cell && cell.vacant) {
        rowHtml += `<td><span class="poste vacant">⚠ Vacant</span></td>`;
      } else {
        rowHtml += "<td></td>";
      }
    });
    tr.innerHTML = rowHtml;
    tbody.appendChild(tr);
  });
  table.appendChild(tbody);
  wrap1.appendChild(table);
  section.appendChild(wrap1);

  // --- Tableau horaire récap ---
  if (DATA[month].hours && DATA[month].hours.length) {
    const h3 = document.createElement("h3");
    h3.className = "section-title";
    h3.textContent = "Tableau horaire";
    section.appendChild(h3);

    const monthHeaders = DATA[month].hours_headers;
    const wrap2 = document.createElement("div");
    wrap2.className = "table-wrap";
    const htable = document.createElement("table");
    htable.className = "hours-table";
    const hthead = document.createElement("thead");
    hthead.innerHTML = "<tr>" + monthHeaders.map(h => `<th>${h}</th>`).join("") + "</tr>";
    htable.appendChild(hthead);
    const htbody = document.createElement("tbody");
    DATA[month].hours.forEach(row => {
      const tr = document.createElement("tr");
      let rowHtml = `<td class="nom">${row.nom}</td>`;
      monthHeaders.slice(1).forEach(h => {
        let v = row[h];
        v = (v === null || v === undefined) ? "" : v;
        let cls = "";
        if (h === "Delta" && typeof v === "number") cls = v < 0 ? "delta-neg" : "delta-pos";
        rowHtml += `<td class="${cls}">${v}</td>`;
      });
      tr.innerHTML = rowHtml;
      htbody.appendChild(tr);
    });
    htable.appendChild(htbody);
    wrap2.appendChild(htable);
    section.appendChild(wrap2);
  }

  container.appendChild(section);
});

function showMonth(month) {
  document.querySelectorAll("#monthNav button").forEach(b => b.classList.toggle("active", b.dataset.month === month));
  document.querySelectorAll(".month").forEach(s => s.classList.toggle("active", s.id === "month-" + month));
}
</script>
</body>
</html>
"""


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 generate_planning_v2.py <fichier.xlsx> [dossier_sortie]")
        sys.exit(1)
    xlsx_path = sys.argv[1]
    out_dir = Path(sys.argv[2]) if len(sys.argv) > 2 else Path(".")
    out_dir.mkdir(parents=True, exist_ok=True)

    data = build_data(xlsx_path)
    html = HTML_TEMPLATE
    html = html.replace("__DATA_JSON__", json.dumps(data, ensure_ascii=False))
    html = html.replace("__HOURS_HEADERS_JSON__", json.dumps([], ensure_ascii=False))
    html = html.replace("__GENERATED_AT__", datetime.now().strftime("%d/%m/%Y %H:%M"))

    out_file = out_dir / "index.html"
    out_file.write_text(html, encoding="utf-8")
    print(f"OK -> {out_file}")


if __name__ == "__main__":
    main()
